"""Report export to XLSX and PDF.

Generated in-memory and streamed back; no temp files. XLSX uses openpyxl,
PDF uses fpdf2 (both pure-Python, no system deps). Kept generic so new report
types are a few lines.
"""

import io
from collections.abc import Sequence
from typing import Any


def to_xlsx(title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet-name limit
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i)].width = 22
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_pdf(title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    from fpdf import FPDF

    pdf = FPDF(orientation="L")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, title, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)

    col_count = len(headers) or 1
    col_width = (pdf.w - 2 * pdf.l_margin) / col_count
    pdf.set_font("Helvetica", "B", 9)
    for header in headers:
        pdf.cell(col_width, 8, str(header)[:30], border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 9)
    for row in rows:
        for value in row:
            pdf.cell(col_width, 7, str(value)[:30], border=1)
        pdf.ln()
    out = pdf.output()
    return bytes(out)
